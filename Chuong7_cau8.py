#Câu 8: Xử lý đọc Excel File
from openpyxl import load_workbook 
wb = load_workbook('demo.xlsx') 
print (wb.sheetnames) 
ws = wb[wb.sheetnames[0]] 
for row in ws.values: 
for value in row: 
print(value,"\t",end='') 
print("")