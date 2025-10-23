#Câu 11: Xử lý Excel File - Viết phần mềm Quản Lý Nhân Viên
print("\n===== CHƯƠNG 7 - CÂU 11 =====")
from openpyxl import Workbook
path = "nhanvien.xlsx"
wb = Workbook()
ws = wb.active
ws.append(["Mã", "Tên", "Tuổi"])
ws.append(["NV01", "Hà", 25])
ws.append(["NV02", "Minh", 22])
ws.append(["NV03", "Lâm", 30])
wb.save(path)
wb = load_workbook(path)
ws = wb.active
data = list(ws.iter_rows(values_only=True))[1:]
data.sort(key=lambda x: x[2])
print("Nhân viên sắp xếp theo tuổi:")
for row in data:
    print(row)